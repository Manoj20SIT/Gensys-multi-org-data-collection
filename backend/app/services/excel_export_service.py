import os
from datetime import datetime
from typing import Any, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelExportError(Exception):
    """Custom exception for excel export failures."""
    pass


class ExcelExportService:
    def __init__(self, export_dir: str = "exports"):
        try:
            self.export_dir = export_dir
            os.makedirs(self.export_dir, exist_ok=True)
        except OSError as e:
            raise ExcelExportError(f"Failed to create export directory '{export_dir}': {e}") from e

    def _flatten_rows(self, run_response: Dict[str, Any]) -> List[Dict[str, Any]]:
        try:
            if not isinstance(run_response, dict):
                raise ValueError("run_response must be a dictionary")

            rows: List[Dict[str, Any]] = []
            results = run_response.get("results", []) or []

            if not isinstance(results, list):
                raise ValueError("'results' must be a list")

            for r in results:
                if not isinstance(r, dict):
                    continue  # skip malformed row safely

                row = {
                    "org_name": r.get("org_name"),
                    # "success": r.get("success"),
                    # "partial_success": r.get("partial_success"),
                    # "expires_in": r.get("expires_in"),
                }

                metrics = r.get("metrics", {}) or {}
                if isinstance(metrics, dict):
                    row.update(metrics)

                rows.append(row)

            return rows

        except Exception as e:
            raise ExcelExportError(f"Failed to flatten rows for excel export: {e}") from e

    def generate_excel(self, run_response: Dict[str, Any]) -> str:
        try:
            rows = self._flatten_rows(run_response)

            wb = Workbook()
            ws = wb.active
            ws.title = "Org Metrics"

            if not rows:
                ws["A1"] = "No data available"
            else:
                headers = list(rows[0].keys())
                ws.append(headers)

                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)

                for col_idx, h in enumerate(headers, start=1):
                    c = ws.cell(row=1, column=col_idx, value=h)
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = Alignment(horizontal="center", vertical="center")

                for row in rows:
                    ws.append([row.get(h, "") for h in headers])

                # autosize columns safely
                for col_idx, h in enumerate(headers, start=1):
                    max_len = len(str(h))
                    for rr in range(2, ws.max_row + 1):
                        val = ws.cell(row=rr, column=col_idx).value
                        if val is not None:
                            max_len = max(max_len, len(str(val)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

                ws.freeze_panes = "A2"

            # UTC now (timezone-safe if your lint checks for it, optionally use datetime.now(timezone.utc))
            file_name = f"org_metrics_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(self.export_dir, file_name)

            try:
                wb.save(file_path)
            except PermissionError as e:
                raise ExcelExportError(f"Permission denied while saving excel file: {file_path}") from e
            except FileNotFoundError as e:
                raise ExcelExportError(f"Export path not found: {file_path}") from e
            except OSError as e:
                raise ExcelExportError(f"OS error while saving excel file '{file_path}': {e}") from e

            return file_name

        except ExcelExportError:
            raise
        except Exception as e:
            raise ExcelExportError(f"Unexpected error while generating excel: {e}") from e
